import openpyxl
import csv
from datetime import datetime

fname = 'people_exs'
file_exs = fname + '.xlsx'

# Створення книги
book = openpyxl.Workbook()

# Видаляємо стандартний аркуш "Sheet", якщо він є
if 'Sheet' in book.sheetnames:
    std = book['Sheet']
    book.remove(std)

# Створення аркушів
blyst1 = book.create_sheet("all", 0)
blyst2 = book.create_sheet("younger_18", 1)
blyst3 = book.create_sheet("18-45", 2)
blyst4 = book.create_sheet("45-70", 3)
blyst5 = book.create_sheet("older_70", 4)

# Створюємо заголовки таблиці
excel_tabl = ['№', 'Прізвище', "Ім'я", 'По-батькові', 'Дата народження', 'Вік']

# Додаємо заголовки на кожен аркуш
blyst1.append(excel_tabl)
blyst2.append(excel_tabl)
blyst3.append(excel_tabl)
blyst4.append(excel_tabl)
blyst5.append(excel_tabl)

def read_csv(filename):
    try:
        # Відкриття CSV файлу
        with open(filename, mode='r', encoding='utf-8') as file:
            reader = csv.DictReader(file)
            index = 1

            # Проходження по кожному рядку CSV
            for row in reader:
                # Визначення віку на основі дати народження
                try:
                    birth = datetime.strptime(row['Дата народження'].split()[0], '%Y-%m-%d')
                except ValueError:
                    print(f"Помилка у форматі дати для рядка {index}: {row['Дата народження']}")
                    continue

                age = datetime.now().year - birth.year - ((datetime.now().month, datetime.now().day) < (birth.month, birth.day))

                # Формування рядка для запису
                salary_seets = [index, row['Прізвище'], row["Ім'я"], row['Побатькові'], row['Дата народження'], age]
                salary_seets_all = [index, row['Прізвище'], row["Ім'я"], row['Побатькові'], row['Стать'], row['Дата народження'], row['Посада'], row['Місто проживання'], row['Адреса'], row['Телефон'], row['Email']]

                # Запис даних на аркуш "all"
                blyst1.append(salary_seets_all)

                # Розподілення по відповідним аркушам залежно від віку
                if age < 18:
                    blyst2.append(salary_seets)
                elif 18 <= age <= 45:
                    blyst3.append(salary_seets)
                elif 45 < age <= 70:
                    blyst4.append(salary_seets)
                else:
                    blyst5.append(salary_seets)

                index += 1

        # Збереження книги
        book.save(file_exs)
        print('ОК')

    except FileNotFoundError:
        print("Помилка: файл CSV не знайдено.")
    except Exception as e:
        print(f"Помилка при відкритті файлу: {e}")

# Виклик функції для читання CSV
read_csv('profeshion.csv')

# Закриття книги
book.close()






# def read_csv(filename):
#     try:
#         with open(filename, "r", encoding='utf-8') as file:
#             reader = csv.reader(file)
#             for row in reader:
#                 print(row)
#     except Exception as e:
#         print(f"Помилка при відкритті файлу: {e}")

# # Виклик функції для зчитування файлу
# read_csv("profeshion.csv")